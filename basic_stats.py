#TODO: Test all functions

import openpyxl
import mpmath
import statistics
import matplotlib
matplotlib.use("TkAgg")
import matplotlib.pyplot as plt
import pylab

from tkinter import *
from tkinter import ttk
from tkinter.filedialog import *


#Normal Model
def invNorm(area):
	#Setting initial result to zero
	result = 0
	#The lower bound is set to -3 because it's unlikely to have something higher than that
	i = -3
	#If the result is not between area-0.0001 and area+0.0001, continue to recalculate the area with the new upper bound value
	while (not((result >= (area-0.000001)) and (result <= (area+0.000001)))):
		i += 0.00001
		result = float(0.5*mpmath.erfc(-((i)/mpmath.sqrt(2))))
	#Return the upper bound value
	return i


#t distribution
def invt(area, df):
	#Setting the initial result to zero
	result = 0
	#Starting at -4 because t values tend to go out further
	i = -4
	#Same as invNorm
	while (not((result >= (area-0.000001)) and (result <= (area+0.000001)))):
		i += 0.00001
		if (i <= 0):
			x2 = (df)/((i**2)+df)
			result = 0.5*mpmath.betainc((df/2), 0.5, 0, x2, regularized=True)
		else:
			x2 = (i**2)/((i**2)+df)
			result = 0.5*(mpmath.betainc(0.5, (df/2), 0, x2, regularized=True)+1)
	return i


def sumData(data, raiseTo):
	dataTemp = []
	for value in data:
		dataTemp.append(value)
	sumOfValues = 0
	for i in range(len(dataTemp)):
		dataTemp[i] = dataTemp[i]**raiseTo
	for value in dataTemp:
		sumOfValues += value
	return sumOfValues


class OneSampleZInterval(object):
	def __init__(self, master):
		#Setting up the main window and central frame
		top = self.top = Toplevel(master)
		self.frame = Frame(top)
		self.frame.grid(column=0, row=0, sticky=(N, W, E, S))
		self.frame.columnconfigure(0, weight=1)
		self.frame.rowconfigure(0, weight=1)
		#Creating Labels for the inputs
		self.successesLabel = Label(self.frame, text='Enter number of successes')
		self.nLabel = Label(self.frame, text='Enter size of sample')
		self.confLabel = Label(self.frame, text='Enter confidence level (0-100)')
		self.successesLabel.grid(row=0, sticky=E)
		self.nLabel.grid(row=1, sticky=E)
		self.confLabel.grid(row=2, sticky=E)
		#Creating input dialogs and gridding them
		self.successesInput = Entry(self.frame)
		self.nInput = Entry(self.frame)
		self.confInput = Entry(self.frame)
		self.successesInput.grid(row=0, column=1)
		self.nInput.grid(row=1, column=1)
		self.confInput.grid(row=2, column=1)
		#Creating Button to retrieve all values
		self.complete = Button(self.frame, text='Submit', command=self.cleanup)
		self.complete.grid(columnspan=2)

	def cleanup(self):
		try:
			self.successesVal = int(self.successesInput.get())
			self.nVal = int(self.nInput.get())
			self.confVal = int(self.confInput.get())
			self.top.destroy()
			self.oneSampleZInterval(self.successesVal, self.nVal, self.confVal)
		except Exception as inst:
			m.result['state'] = 'normal'
			m.result.insert(1.0, inst)
			m.result['state'] = 'disabled'

	def oneSampleZInterval(self, successes, n, confidenceLevel):
		try:
			#Setting the output to normal then clearing it
			m.result['state'] = 'normal'
			m.result.delete(1.0, END)
			#Calculating p from the sample
			self.pHat = successes/n
			#Getting Critical value
			self.zCritical = -1*invNorm((1-(confidenceLevel/100))/2)
			#Getting Standard deviation
			self.stDev = float(mpmath.sqrt(((self.pHat)*(1-self.pHat))/(n)))
			self.marError = self.stDev*self.zCritical
			self.CLower = self.pHat - self.marError
			self.CUpper = self.pHat + self.marError
			m.result.insert(1.0, 'p-Hat = ' + str(self.pHat) + '\n')
			m.result.insert(2.0, 'StDev = ' + str(self.stDev) + '\n')
			m.result.insert(3.0, 'ME = ' + str(self.marError) + '\n')
			m.result.insert(4.0, 'CLower = ' + str(self.CLower) + '\n')
			m.result.insert(5.0, 'CUpper = ' + str(self.CUpper))
			m.result['state'] = 'disabled'
		except Exception as inst:
			m.result.insert(1.0, inst)


class TwoSampleZInterval(object):
	def __init__(self, master):
		#Setting up the main window and central frame
		top = self.top = Toplevel(master)
		self.frame = Frame(top)
		self.frame.grid(column=0, row=0, sticky=(N, W, E, S))
		self.frame.columnconfigure(0, weight=1)
		self.frame.rowconfigure(0, weight=1)
		#Creating Labels for the inputs
		self.successesLabel1 = Label(self.frame, text='Enter number of successes for sample 1')
		self.nLabel1 = Label(self.frame, text='Enter size of sample 1')
		self.successesLabel2 = Label(self.frame, text='Enter number of successes for sample 2')
		self.nLabel2 = Label(self.frame, text='Enter size of sample 2')
		self.confLabel = Label(self.frame, text='Enter confidence level (0-100)')
		self.successesLabel1.grid(row=0, sticky=E)
		self.nLabel1.grid(row=1, sticky=E)
		self.successesLabel2.grid(row=2, sticky=E)
		self.nLabel2.grid(row=3, sticky=E)
		self.confLabel.grid(row=4, sticky=E)
		#Creating input dialogs and gridding them
		self.successesInput1 = Entry(self.frame)
		self.nInput1 = Entry(self.frame)
		self.successesInput2 = Entry(self.frame)
		self.nInput2 = Entry(self.frame)
		self.confInput = Entry(self.frame)
		self.successesInput1.grid(row=0, column=1)
		self.nInput1.grid(row=1, column=1)
		self.successesInput2.grid(row=2, column=1)
		self.nInput2.grid(row=3, column=1)
		self.confInput.grid(row=4, column=1)
		#Creating Button to retrieve all values
		self.complete = Button(self.frame, text='Submit', command=self.cleanup)
		self.complete.grid(row=5, column=1)

	def cleanup(self):
		try:
			self.successesVal1 = int(self.successesInput1.get())
			self.nVal1 = int(self.nInput1.get())
			self.successesVal2 = int(self.successesInput2.get())
			self.nVal2 = int(self.nInput2.get())
			self.confVal = int(self.confInput.get())
			self.top.destroy()
			self.twoSampleZInterval(self.successesVal1, self.nVal1, self.successesVal2, self.nVal2, self.confVal)
		except Exception as inst:
			m.result['state'] = 'normal'
			m.result.insert(1.0, inst)
			m.result['state'] = 'disabled'

	def twoSampleZInterval(self, successes1, n1, successes2, n2, confidenceLevel):
		try:
			#Setting the output to normal then clearing it
			m.result['state'] = 'normal'
			m.result.delete(1.0, END)
			#Calculating p from the samples
			self.pHat1 = successes1/n1
			self.pHat2 = successes2/n2
			#Getting the difference in p
			self.pDiff = self.pHat1-self.pHat2
			#Getting critical value
			self.zCritical = -1*invNorm((1-(confidenceLevel/100))/2)
			#Getting Standard error
			self.stDev = mpmath.sqrt((((self.pHat1)*(1-self.pHat1))/n1) + (((self.pHat2)*(1-self.pHat2))/n2))
			self.marError = self.zCritical*self.stDev
			self.CLower = self.pDiff - self.marError
			self.CUpper = self.pDiff + self.marError
			m.result.insert(1.0, 'p-Hat1 = ' + str(self.pHat1) + '\n')
			m.result.insert(2.0, 'pHat2 = ' + str(self.pHat2) + '\n')
			m.result.insert(3.0, 'pDiff = ' + str(self.pDiff) + '\n')
			m.result.insert(4.0, 'ME  = ' + str(self.pDiff) + '\n')
			m.result.insert(5.0, 'Clower = ' + str(self.CLower) + '\n')
			m.result.insert(6.0, 'CUpper = ' + str(self.CUpper))
			m.result['state'] = 'disabled'
		except Exception as inst:
			m.result.insert(1.0, inst)

#Confidence Intervals for means; reading from an excel file
class OneSampleTInterval(object):
	def __init__(self, master):
		#Setting up the main window and central frame
		top = self.top = Toplevel(master)
		self.frame = Frame(top)
		self.frame.grid(column=0, row=0, sticky=(N, W, E, S))
		self.frame.columnconfigure(0, weight=1)
		self.frame.rowconfigure(0, weight=1)
		#Creating Labels for the inputs
		self.sheetLabel = Label(self.frame, text='Enter the name of the sheet')
		self.columnLabel1 = Label(self.frame, text='Enter column number')
		self.confLabel = Label(self.frame, text='Enter confidence level (0-100)')
		self.sheetLabel.grid(row=0, sticky=E)
		self.columnLabel1.grid(row=1, sticky=E)
		self.confLabel.grid(row=2, sticky=E)
		#Creating inputs
		self.srcVAL = filedialog.askopenfilename(title="Locate Excel file", filetypes=(("excel files", "*.xlsx"), ("all files", "*.*")))
		self.sheetVal = Entry(self.frame)
		self.sheetVal.grid(row=0, column=1)
		self.columnVal1 = Entry(self.frame)
		self.columnVal1.grid(row=1, column=1)
		self.confVal = Entry(self.frame)
		self.confVal.grid(row=2, column=1)
		#Creating Button to retrieve all values
		self.complete = Button(self.frame, text='Submit', command=self.cleanup)
		self.complete.grid(row=3, column=1)
	def cleanup(self):
		try:
			self.source = self.srcVAL
			self.sheetName = self.sheetVal.get()
			self.columnNum = int(self.columnVal1.get())
			self.confLevel = int(self.confVal.get())
			self.top.destroy()
			self.oneSampleTInterval(self.source, self.sheetName, self.confLevel, self.columnNum)
		except Exception as inst:
			m.result['state'] = 'normal'
			m.result.insert(1.0, inst)
			m.result['state'] = 'disabled'

	def oneSampleTInterval(self, src, setSheet, confidenceLevel, column):
		try:
			#Setting the output to normal then clearing it
			m.result['state'] = 'normal'
			m.result.delete(1.0, END)
			#Opening the workbook and setting the sheet
			self.wb = openpyxl.load_workbook(src)
			self.sheet = self.wb.get_sheet_by_name(setSheet)
			#Getting a list of the cells in order to calculate length
			self.dataList = list(self.sheet.columns)[column]
			#Populating a list with data from the cells
			self.data = []
			for cellObj in self.dataList:
				self.data.append(float(cellObj.value))
			#Getting mean,Standard Deviation,n and df
			self.mean = statistics.mean(self.data)
			self.sampleStDev = statistics.stdev(self.data)
			self.n = len(self.dataList)
			self.df = self.n-1
			self.stDev = (self.sampleStDev)/(mpmath.sqrt(self.n))
			#Getting critical value
			self.tCritical = -1*invt(((1-(confidenceLevel/100))/2), self.df)
			self.stError = self.stDev*self.tCritical
			self.CLower = self.mean - self.stError
			self.CUpper = self.mean + self.stError
			m.result.insert(1.0, 'Mean = ' + str(self.mean) + '\n')
			m.result.insert(2.0, 'Standard Deviation = ' + str(self.stDev) + '\n')
			m.result.insert(3.0, 'Critical Value = ' + str(self.tCritical) + '\n')
			m.result.insert(4.0, 'Standard Error = ' + str(self.stError) + '\n')
			m.result.insert(5.0, 'Lower limit = ' + str(self.CLower) + '\n')
			m.result.insert(6.0, 'Upper limit = ' + str(self.CUpper) + '\n')
			m.result['state'] = 'disabled'
		except FileNotFoundError:
			m.result.insert(1.0, 'File not Found')
		except KeyError:
			m.result.insert(1.0, 'Sheet does not exist')
		except TypeError:
			m.result.insert(1.0, 'File contains letters or empty cells')
		except Exception as inst:
			m.result.insert(1.0, inst)

class TwoSampleTInterval(object):
	def __init__(self, master):
		#Setting up the main window and central frame
		top = self.top = Toplevel(master)
		self.frame = Frame(top)
		self.frame.grid(column=0, row=0, sticky=(N, W, E, S))
		self.frame.columnconfigure(0, weight=1)
		self.frame.rowconfigure(0, weight=1)
		#Creating Labels for the inputs
		self.sheetLabel = Label(self.frame, text='Enter the name of the sheet')
		self.columnLabel1 = Label(self.frame, text='Enter column number 1')
		self.columnLabel2 = Label(self.frame, text='Enter column number 2')
		self.confLabel = Label(self.frame, text='Enter confidence level (0-100)')
		self.sheetLabel.grid(row=0, sticky=E)
		self.columnLabel1.grid(row=1, sticky=E)
		self.columnLabel1.grid(row=2, sticky=E)
		self.confLabel.grid(row=3, sticky=E)
		#Creating inputs
		self.srcVAL = filedialog.askopenfilename(title="Locate Excel file", filetypes=(("excel files", "*.xlsx"), ("all files", "*.*")))
		self.sheetVal = Entry(self.frame)
		self.sheetVal.grid(row=0, column=1)
		self.columnVal1 = Entry(self.frame)
		self.columnVal1.grid(row=1, column=1)
		self.columnVal2 = Entry(self.frame)
		self.columnVal2.grid(row=2, column=1)
		self.confVal = Entry(self.frame)
		self.confVal.grid(row=3, column=1)
		#Creating Button to retrieve all values
		self.complete = Button(self.frame, text='Submit', command=self.cleanup)
		self.complete.grid(row=4, column=1)

	def cleanup(self):
		try:
			self.source = self.srcVAL
			self.sheetName = self.sheetVal.get()
			self.columnNum1 = int(self.columnVal1.get())
			self.columnNum2 = int(self.columnVal2.get())
			self.confLevel = int(self.confVal.get())
			self.top.destroy()
			self.twoSampleTInterval(self.source, self.sheetName, self.confLevel, self.columnNum1, self.columnNum2)
		except Exception as inst:
			m.result['state'] = 'normal'
			m.result.insert(1.0, inst)
			m.result['state'] = 'disabled'

	def twoSampleTInterval(self, src, setSheet, confidenceLevel, column1, column2):
		try:
			#Opening the workbook and setting the sheet
			self.wb = openpyxl.load_workbook(src)
			self.sheet = self.wb.get_sheet_by_name(setSheet)
			#Getting a list of the cells in order to calculate length
			self.dataList1 = list(self.sheet.columns)[column1]
			self.dataList2 = list(self.sheet.columns)[column2]
			#Populating a list with data from the cells
			self.data1, self.data2 = [], []
			for cellObj in self.dataList1:
				self.data1.append(float(cellObj.value))
			for cellObj in self.dataList2:
				self.data2.append(float(cellObj.value))
			#Getting mean, Standard Deviation, n and df for both samples
			self.mean1 = statistics.mean(self.data1)
			self.sampleStDev1 = statistics.stdev(self.data1)
			self.n1 = len(self.dataList1)
			self.df1 = self.n1-1
			self.mean2 = statistics.mean(self.data2)
			self.sampleStDev2 = statistics.stdev(self.data2)
			self.n2 = len(self.dataList2)
			self.df2 = self.n2-1
			self.meanDiff = self.mean1-self.mean2
			#BUG: The df calculation is wrong
			self.df = (self.n1+self.n2)-2
			self.stDev = float(mpmath.sqrt(((self.sampleStDev1**2)/self.n1)+((self.sampleStDev2**2)/self.n2)))
			#Getting critical value
			self.tCritical = -1*invt(((1-(confidenceLevel/100))/2), self.df)
			self.stError = self.stDev*self.tCritical
			self.CLower = self.meanDiff - self.stError
			self.CUpper = self.meanDiff + self.stError
			m.result.insert(1.0, 'Clower = ' + str(self.CLower) + '\n')
			m.result.insert(2.0, 'CUpper = ' + str(self.CUpper) + '\n')
			m.result.insert(3.0, 'Difference in mean = ' + str(self.meanDiff) + '\n')
			m.result.insert(4.0, 'SE = ' + str(self.stError) + '\n')
			m.result.insert(5.0, 'df = ' + str(self.df) + '\n')
			m.result.insert(6.0, 'mean1 = ' + str(self.mean1) + '\n')
			m.result.insert(7.0, 'mean2 = ' + str(self.mean2) + '\n')
			m.result.insert(8.0, 'stDev1 = ' + str(self.sampleStDev1) + '\n')
			m.result.insert(9.0, 'stDev2 = ' + str(self.sampleStDev2) + '\n')
			m.result.insert(10.0, 'n1 = ' + str(self.n1) + '\n')
			m.result.insert(11.0, 'n2 = ' + str(self.n2) + '\n')
			m.result['state'] = 'disabled'
		except FileNotFoundError:
			m.result.insert(1.0, 'File not Found')
		except KeyError:
			m.result.insert(1.0, 'Sheet does not exist')
		except TypeError:
			m.result.insert(1.0, 'File contains letters or empty cells')
		except Exception as inst:
			m.result.insert(1.0, inst)


class OneVarStats(object):
	def __init__(self, master):
		#Setting up the main window and central frame
		top = self.top = Toplevel(master)
		self.frame = Frame(top)
		self.frame.grid(column=0, row=0, sticky=(N, W, E, S))
		self.frame.columnconfigure(0, weight=1)
		self.frame.rowconfigure(0, weight=1)
		#Creating Labels for the inputs
		self.sheetLabel = Label(self.frame, text='Enter the name of the sheet')
		self.columnLabel = Label(self.frame, text='Enter column number')
		self.sheetLabel.grid(row=0, sticky=E)
		self.columnLabel.grid(row=1, sticky=E)
		#Creating inputs
		self.srcVAL = filedialog.askopenfilename(title="Locate Excel file", filetypes=(("excel files", "*.xlsx"), ("all files", "*.*")))
		self.sheetVal = Entry(self.frame)
		self.sheetVal.grid(row=0, column=1)
		self.columnVal = Entry(self.frame)
		self.columnVal.grid(row=1, column=1)
		#Creating Button to retrieve all values
		self.complete = Button(self.frame, text='Submit', command=self.cleanup)
		self.complete.grid(row=2, column=1)

	def cleanup(self):
		try:
			self.source = self.srcVAL
			self.sheetName = self.sheetVal.get()
			self.column = int(self.columnVal.get())
			self.top.destroy()
			self.oneVarStats(self.source, self.sheetName, self.column)
		except Exception as inst:
			m.result['state'] = 'normal'
			m.result.insert(1.0, inst)
			m.result['state'] = 'disabled'

	def oneVarStats(self, src, setSheet, column):
		try:
			#Setting the output to normal then clearing it
			m.result['state'] = 'normal'
			m.result.delete(1.0, END)
			#Opening the workbook and setting the sheet
			self.wb = openpyxl.load_workbook(src)
			self.sheet = self.wb.get_sheet_by_name(setSheet)
			#Getting a list of the cells in order to calculate length
			self.dataList = list(self.sheet.columns)[column]
			#Populating a list with data from the cells
			self.data = []
			for cellObj in self.dataList:
				self.data.append(float(cellObj.value))
			self.mean = statistics.mean(self.data)
			self.sampleStDev = statistics.stdev(self.data)
			self.populationStDev = statistics.pstdev(self.data)
			self.n = len(self.data)
			self.sumOfValues = sumData(self.data, 1)
			self.sumOfValuesSquared = sumData(self.data, 2)
			#Five Number summary
			self.data.sort()
			self.minVal = self.data[0]
			self.q1Val, self.medianVal, self.q3Val = 0, 0, 0
			self.maxVal = self.data[len(self.data)-1]
			#If the data has odd number of values
			if (len(self.data) % 2 != 0):
				self.middleIndex = (len(self.data)-1)//2
				self.medianVal = self.data[self.middleIndex]
				self.lowerData = self.data[0:self.middleIndex]
				self.q1Val = statistics.median(self.lowerData)
				self.upperData = self.data[self.middleIndex+1:len(self.data)]
				self.q3Val = statistics.median(self.upperData)
			else:
				self.medianVal = statistics.median(self.data)
				self.lowerData = self.data[0:len(self.data)//2]
				self.q1Val = statistics.median(self.lowerData)
				self.upperData = self.data[len(self.data)//2:len(self.data)]
				self.q3Val = statistics.median(self.upperData)
			#Sum of squared deviations
			self.ssx = 0
			for value in self.data:
				self.ssx += (value-self.mean)**2
			#Histogram and Boxplot
			self.fig = plt.figure()
			self.ax = self.fig.add_subplot(111)
			self.ax.hist(self.data)
			self.fig2 = plt.figure()
			self.ax2 = self.fig2.add_subplot(111)
			self.ax2.boxplot(self.data)
			#Appending results to output box
			m.result.insert(1.0, 'Mean = ' + str(self.mean) + '\n')
			m.result.insert(2.0, 'Sum of x = ' + str(self.sumOfValues) + '\n')
			m.result.insert(3.0, 'Sum of squared x = ' + str(self.sumOfValuesSquared) + '\n')
			m.result.insert(4.0, 'Sample standard deviation = ' + str(self.sampleStDev) + '\n')
			m.result.insert(5.0, 'Population standard deviation = ' + str(self.populationStDev) + '\n')
			m.result.insert(6.0, 'n = ' + str(self.n) + '\n')
			m.result.insert(7.0, 'Min = ' + str(self.minVal) + '\n')
			m.result.insert(8.0, 'Q1 = ' + str(self.q1Val) + '\n')
			m.result.insert(9.0, 'Median = ' + str(self.medianVal) + '\n')
			m.result.insert(10.0, 'Q3 = ' + str(self.q3Val) + '\n')
			m.result.insert(11.0, 'Max = ' + str(self.maxVal) + '\n')
			m.result.insert(12.0, 'Sum of squared deviations = ' + str(self.ssx) + '\n')
			m.result['state'] = 'disabled'
			plt.show()
		except FileNotFoundError:
			m.result.insert(1.0, 'File not found')
		except KeyError:
			m.result.insert(1.0, 'Sheet does not exist')
		except TypeError:
			m.result.insert(1.0, 'File contains letters or empty cells')
		except Exception as inst:
			m.result.insert(1.0, inst)


class LinReg(object):
	def __init__(self, master):
		#Setting up the main window and central frame
		top = self.top = Toplevel(master)
		self.frame = Frame(top)
		self.frame.grid(column=0, row=0, sticky=(N, W, E, S))
		self.frame.columnconfigure(0, weight=1)
		self.frame.rowconfigure(0, weight=1)
		#Creating Labels for the inputs
		self.sheetLabel = Label(self.frame, text='Enter the name of the sheet')
		self.columnLabel1 = Label(self.frame, text='Enter column number 1')
		self.columnLabel2 = Label(self.frame, text='Enter column number 2')
		self.sheetLabel.grid(row=0, sticky=E)
		self.columnLabel1.grid(row=1, sticky=E)
		self.columnLabel2.grid(row=2, sticky=E)
		#Creating inputs
		self.srcVAL = filedialog.askopenfilename(title="Locate Excel file", filetypes=(("excel files", "*.xlsx"), ("all files", "*.*")))
		self.sheetVal = Entry(self.frame)
		self.sheetVal.grid(row=0, column=1)
		self.columnVal1 = Entry(self.frame)
		self.columnVal1.grid(row=1, column=1)
		self.columnVal2 = Entry(self.frame)
		self.columnVal2.grid(row=2, column=1)
		#Creating Button to retrieve all values
		self.complete = Button(self.frame, text='Submit', command=self.cleanup)
		self.complete.grid(row=3, column=1)

	def cleanup(self):
		try:
			self.source = self.srcVAL
			self.sheetName = self.sheetVal.get()
			self.column1 = int(self.columnVal1.get())
			self.column2 = int(self.columnVal2.get())
			self.top.destroy()
			self.linReg(self.source, self.sheetName, self.column1, self.column2)
		except Exception as inst:
			m.result['state'] = 'normal'
			m.result.insert(1.0, inst)
			m.result['state'] = 'disabled'

	def linReg(self, src, setSheet, column1, column2):
		try:
			#Setting the output to normal then clearing it
			m.result['state'] = 'normal'
			m.result.delete(1.0, END)
			#Opening the workbook and setting the sheet
			self.wb = openpyxl.load_workbook(src)
			self.sheet = self.wb.get_sheet_by_name(setSheet)
			#Getting a list of the cells in order to calculate length
			self.dataList1 = list(self.sheet.columns)[column1]
			self.dataList2 = list(self.sheet.columns)[column2]
			#Populating a list with data from the cells
			self.data1, self.data2 = [], []
			for cellObj in self.dataList1:
				self.data1.append(float(cellObj.value))
			for cellObj in self.dataList2:
				self.data2.append(float(cellObj.value))
			if (len(self.data1) != len(self.data2)):
				raise Exception('Data is not of equal n')
			(self.m, self.b) = pylab.polyfit(self.data1, self.data2, 1)
			self.yp = pylab.polyval([self.m, self.b], self.data1)
			self.mean1 = statistics.mean(self.data1)
			self.mean2 = statistics.mean(self.data2)
			self.stDev1 = statistics.stdev(self.data1)
			self.stDev2 = statistics.stdev(self.data2)
			self.r = (self.m*self.stDev1)/self.stDev2
			self.rSquared = self.r**2
			m.result.insert(1.0, 'a = ' + str(self.b) + '\n')
			m.result.insert(2.0, 'b = ' + str(self.m) + '\n')
			m.result.insert(3.0, 'r = ' + str(self.r) + '\n')
			m.result.insert(4.0, 'r-squared = ' + str(self.rSquared) + '\n')
			m.result['state'] = 'disabled'
			self.fig = plt.figure()
			self.ax = self.fig.add_subplot(111)
			self.ax.plot(self.data1, self.data2, 'ro')
			self.ax.plot(self.data1, self.yp)
			plt.show()
		except FileNotFoundError:
			m.result.insert(1.0, 'File not found')
		except KeyError:
			m.result.insert(1.0, 'Sheet does not exist')
		except TypeError:
			m.result.insert(1.0, 'File contains letters or empty cells')
		except Exception as inst:
			m.result.insert(1.0, inst)


class main(object):
	def __init__(self, master):
		self.master = master
		# *******Menu Bar********
		self.menubar = Menu(master)
		self.master.config(menu=self.menubar)
		#Adding stats calculations sub-menu
		self.statCalc = Menu(self.menubar, tearoff=0)
		self.menubar.add_cascade(label='Stat Calculations', menu=self.statCalc)
		self.statCalc.add_command(label='One Variable Statistics', command=self.initOneVarStats)
		self.statCalc.add_command(label='Linear Regression', command=self.initLinReg)
		#Adding confidence intervals sub-menu
		self.confInt = Menu(self.menubar, tearoff=0)
		self.menubar.add_cascade(label='Confidence Intervals', menu=self.confInt)
		self.confInt.add_command(label='One Sample Z Interval', command=self.initOneSampleZInterval)
		self.confInt.add_command(label='Two Sample Z Interval', command=self.initTwoSampleZInterval)
		self.confInt.add_separator()
		self.confInt.add_command(label='One Sample T Interval', command=self.initOneSampleTInterval)
		self.confInt.add_command(label='Two Sample T Interval', command=self.initTwoSampleTInterval)
		# *******output area********
		self.resultLabel = ttk.Labelframe(master, text='Output')
		self.result = Text(self.resultLabel, state='disabled')
		self.result.insert('1.0', 'Result will be shown here')
		self.resultLabel.pack()
		self.result.pack()

	def initOneSampleZInterval(self):
		self.w = OneSampleZInterval(self.master)
		self.master.wait_window(self.w.top)

	def initTwoSampleZInterval(self):
		self.w = TwoSampleZInterval(self.master)
		self.master.wait_window(self.w.top)

	def initOneVarStats(self):
		self.w = OneVarStats(self.master)
		self.master.wait_window(self.w.top)

	def initLinReg(self):
		self.w = LinReg(self.master)
		self.master.wait_window(self.w.top)

	def initOneSampleTInterval(self):
		self.w = OneSampleTInterval(self.master)
		self.master.wait_window(self.w.top)

	def initTwoSampleTInterval(self):
		self.w = TwoSampleTInterval(self.master)
		self.master.wait_window(self.w.top)


# *******Main Window********
root = Tk()
root.title('Basic Statistics')
root.geometry('400x300')
m = main(root)
root.mainloop()
